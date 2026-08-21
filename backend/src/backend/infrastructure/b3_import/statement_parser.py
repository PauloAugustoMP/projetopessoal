"""Parser for the B3 investor-portal "extrato de movimentação" (CSV or Excel).

Expected columns (as exported by the portal):
    Entrada/Saída; Data; Movimentação; Produto; Instituição; Quantidade;
    Preço unitário; Valor da Operação

Each row is classified as a trade (buy/sell), a dividend, or a corporate action
(docs/business-rules.md §7). Known-irrelevant movement types are skipped
silently; anything unrecognized or missing required data lands in the
manual-review list instead of failing the whole import.

Corporate-action quantity semantics (how the portal reports them):
- Desdobro (split) / Bonificação: a CREDIT with the ADDITIONAL shares received.
- Grupamento (reverse split): a DEBIT with the shares removed, or a CREDIT with
  the new total — both are accepted; the factor is derived later against the
  position on that date.
"""

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Literal, Optional


@dataclass(frozen=True)
class ParsedTrade:
    row: int
    ticker: str
    name: str
    type: Literal["buy", "sell"]
    quantity: float
    price_per_share: float
    date: str  # ISO


@dataclass(frozen=True)
class ParsedDividend:
    row: int
    ticker: str
    name: str
    type: Literal["dividend", "jcp", "reit_income"]
    quantity: float
    gross_value_per_share: float
    payment_date: str  # ISO


@dataclass(frozen=True)
class ParsedCorporateActionEvent:
    row: int
    ticker: str
    name: str
    type: Literal["split", "reverse_split", "bonus_shares"]
    date: str  # ISO
    direction: Literal["credit", "debit"]
    quantity: float


@dataclass(frozen=True)
class ReviewRow:
    row: int
    reason: str
    raw: str


@dataclass
class ParsedStatement:
    trades: list[ParsedTrade] = field(default_factory=list)
    dividends: list[ParsedDividend] = field(default_factory=list)
    corporate_actions: list[ParsedCorporateActionEvent] = field(default_factory=list)
    review_rows: list[ReviewRow] = field(default_factory=list)


class StatementFormatError(Exception):
    pass


def _normalize(text: str) -> str:
    stripped = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return " ".join(stripped.lower().split())


HEADER_SEARCH_ROWS = 15

EXPECTED_COLUMNS = {
    "entrada/saida": "direction",
    "data": "date",
    "movimentacao": "movement",
    "produto": "product",
    "instituicao": "institution",
    "quantidade": "quantity",
    "preco unitario": "unit_price",
    "valor da operacao": "total_value",
}

# "Transferência - Liquidação" is the portal's own wording; brokers that feed the
# statement sometimes use plain "Compra"/"Venda" for the very same settlement.
TRADE_MOVEMENTS = {"transferencia - liquidacao", "compra", "venda"}
EXPLICIT_SELL_MOVEMENTS = {"venda"}
EXPLICIT_BUY_MOVEMENTS = {"compra"}
SUBSCRIPTION_BUY_MOVEMENTS = {
    "direitos de subscricao - exercido",
    "direito de subscricao - exercido",
    "subscricao exercida",
}
DIVIDEND_MOVEMENTS: dict[str, Literal["dividend", "jcp", "reit_income"]] = {
    "dividendo": "dividend",
    "juros sobre capital proprio": "jcp",
    "rendimento": "reit_income",
}
CORPORATE_ACTION_MOVEMENTS: dict[str, Literal["split", "reverse_split", "bonus_shares"]] = {
    "desdobro": "split",
    "desdobramento": "split",
    "grupamento": "reverse_split",
    "bonificacao em ativos": "bonus_shares",
}
IGNORED_MOVEMENTS = {
    "atualizacao",
    # Subscription lifecycle: receiving, requesting or letting rights lapse does
    # not move the position. Only an *exercised* right does, and it arrives as its
    # own row (SUBSCRIPTION_BUY_MOVEMENTS) — docs/business-rules.md §5.
    "direito de subscricao",
    "direitos de subscricao",
    "direito sobras de subscricao",
    "direito sobras de subscricao - nao exercido",
    "direitos sobras de subscricao",
    "solicitacao de subscricao",
    "recibo de subscricao",
    "cessao de direitos - direitos de subscricao",
    "emprestimo",
    "emprestimo de ativos",
    "transferencia",
    "direitos de subscricao - nao exercido",
    "direito de subscricao - nao exercido",
    "cessao de direitos",
    "cessao de direitos - solicitada",
    "leilao de fracao",
    "fracao em ativos",
}


def _parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in ("", "-"):
        return None
    text = text.replace("R$", "").strip()
    # Brazilian format: 1.234,56 — and 1.000 (dot as thousands separator, no decimals)
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(\.\d{3})+", text):
        text = text.replace(".", "")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(value: object) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# B3 tickers are letters followed by a numeric suffix: ITSA4, TAEE11, AFHI13,
# BOVA11, AAPL34. Requiring the digits is what separates a real ticker from a
# plain description ("TESOURO SELIC 2029", "CDB BANCO X"), which belongs to the
# review queue rather than the catalog.
TICKER_PATTERN = re.compile(r"^[A-Z]{4,6}\d{1,3}$")


def _extract_product(product: object) -> tuple[Optional[str], str]:
    """The Produto column normally reads "ITSA4 - ITAUSA S.A.": ticker, separator,
    name. Not every row follows it — some carry only a description, and taking the
    whole string as a ticker would push a 30-character value into a 20-character
    column. So the candidate has to actually look like a ticker; when it does not,
    the row goes to manual review instead."""
    text = str(product or "").strip()
    if not text:
        return None, ""

    head, separator, name = text.partition(" - ")
    candidate = head.strip().upper()
    if TICKER_PATTERN.match(candidate):
        return candidate, (name.strip() or candidate)

    # No " - " separator, or something else in front of it: fall back to the first
    # whitespace-delimited token, which is where the ticker sits in every layout
    # seen so far ("ITSA4 ITAUSA S.A.").
    first_token = text.split()[0].strip().upper() if text.split() else ""
    if TICKER_PATTERN.match(first_token):
        remainder = text[len(first_token) :].lstrip(" -").strip()
        return first_token, (remainder or first_token)

    return None, (text if not separator else text)


def _read_raw_rows(content: bytes, filename: str) -> list[list[object]]:
    if filename.lower().endswith((".xlsx", ".xls")):
        return _read_excel_rows(content)

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=_detect_delimiter(text))]


def _header_score(rows: list[list[object]]) -> int:
    """How many expected columns the best candidate header row recognizes."""
    return max(
        (
            sum(1 for cell in row if _normalize(str(cell or "")) in EXPECTED_COLUMNS)
            for row in rows[:HEADER_SEARCH_ROWS]
        ),
        default=0,
    )


def _read_excel_rows(content: bytes) -> list[list[object]]:
    """Reads the sheet that actually holds the statement.

    Two things make this less trivial than "read the active sheet": exports can
    carry several sheets (cover page, filters, data), and openpyxl's read_only
    mode trusts the file's declared dimensions, which some exporters get wrong —
    yielding truncated rows. So we score every sheet by how many header columns
    it recognizes, and fall back to the full (slower, more forgiving) parser when
    the fast path finds nothing usable.
    """
    from openpyxl import load_workbook

    best: list[list[object]] = []
    for read_only in (True, False):
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=read_only, data_only=True)
        except Exception as error:
            raise StatementFormatError(
                f"Could not read the Excel file: {error}. If it came from an older "
                "'.xls' export, re-save it as .xlsx or CSV."
            ) from error

        for sheet in workbook.worksheets:
            if read_only:
                # Ignore the declared dimensions; scan what the sheet really holds.
                sheet.reset_dimensions()
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if _header_score(rows) > _header_score(best):
                best = rows

        workbook.close()
        if _header_score(best) > 1:
            break

    return best


def _detect_delimiter(text: str) -> str:
    """The portal exports semicolon CSV, but "save as" round-trips through Excel
    produce tab- or comma-separated files just as often. Picking by raw frequency
    is wrong — decimal commas ("9,10") outnumber real separators — so we score each
    candidate by how well it splits a header row into recognizable columns. The
    search spans the first rows because exports may carry title lines on top."""
    head = "\n".join(text.splitlines()[:HEADER_SEARCH_ROWS])

    def score(delimiter: str) -> tuple[int, int]:
        rows = list(csv.reader(io.StringIO(head), delimiter=delimiter))
        best_recognized = max(
            (sum(1 for cell in row if _normalize(str(cell)) in EXPECTED_COLUMNS) for row in rows),
            default=0,
        )
        widest = max((len(row) for row in rows), default=0)
        return best_recognized, widest

    return max(("\t", ";", ",", "|"), key=score)


def _locate_header(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    """Finds the header row: exports sometimes carry title/filter lines above it,
    so we take the row in the first few that recognizes the most columns."""
    best_index, best_columns = 0, {}
    for index, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        labels = [_normalize(str(cell or "")) for cell in row]
        columns = {
            EXPECTED_COLUMNS[label]: position
            for position, label in enumerate(labels)
            if label in EXPECTED_COLUMNS
        }
        if len(columns) > len(best_columns):
            best_index, best_columns = index, columns
    return best_index, best_columns


def parse_statement(content: bytes, filename: str) -> ParsedStatement:
    raw_rows = _read_raw_rows(content, filename)
    if not raw_rows:
        raise StatementFormatError("The file is empty.")

    header_index, columns = _locate_header(raw_rows)
    missing = set(EXPECTED_COLUMNS.values()) - {"institution", "total_value"} - set(columns)
    if missing:
        found = [
            str(cell or "").strip()
            for cell in (raw_rows[header_index] if header_index < len(raw_rows) else [])
            if str(cell or "").strip()
        ]
        raise StatementFormatError(
            "The file does not look like a B3 movement statement — missing columns: "
            + ", ".join(sorted(missing))
            + f". Columns found: {found or 'none'}. Export the 'Movimentação' statement "
            "from the B3 investor portal without editing it."
        )

    result = ParsedStatement()

    for row_number, cells in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        if not any(str(cell or "").strip() for cell in cells):
            continue
        raw = "; ".join(str(cell or "") for cell in cells)

        def cell(name: str) -> object:
            index = columns.get(name)
            return cells[index] if index is not None and index < len(cells) else None

        movement = _normalize(str(cell("movement") or ""))
        direction_text = _normalize(str(cell("direction") or ""))
        direction = "credit" if direction_text.startswith("cred") else "debit"
        ticker, product_name = _extract_product(cell("product"))
        parsed_date = _parse_date(cell("date"))
        quantity = _parse_number(cell("quantity"))
        unit_price = _parse_number(cell("unit_price"))

        def review(reason: str) -> None:
            result.review_rows.append(ReviewRow(row=row_number, reason=reason, raw=raw))

        if movement in IGNORED_MOVEMENTS:
            continue
        if ticker is None or parsed_date is None:
            review(
                f'Could not read a ticker from the Produto column ("{product_name[:80]}") '
                "or the date is unreadable."
                if parsed_date is not None
                else "Missing or unreadable ticker/date."
            )
            continue

        if movement in TRADE_MOVEMENTS or movement in SUBSCRIPTION_BUY_MOVEMENTS:
            if quantity is None or quantity <= 0 or unit_price is None:
                review("Trade row without a readable quantity/unit price.")
                continue
            if movement in EXPLICIT_BUY_MOVEMENTS or movement in SUBSCRIPTION_BUY_MOVEMENTS:
                trade_type: Literal["buy", "sell"] = "buy"
            elif movement in EXPLICIT_SELL_MOVEMENTS:
                trade_type = "sell"
            else:
                trade_type = "buy" if direction == "credit" else "sell"
            result.trades.append(
                ParsedTrade(
                    row=row_number,
                    ticker=ticker,
                    name=product_name,
                    type=trade_type,
                    quantity=quantity,
                    price_per_share=unit_price,
                    date=parsed_date,
                )
            )
        elif movement in DIVIDEND_MOVEMENTS:
            if quantity is None or quantity <= 0 or unit_price is None:
                review("Dividend row without a readable quantity/unit value.")
                continue
            result.dividends.append(
                ParsedDividend(
                    row=row_number,
                    ticker=ticker,
                    name=product_name,
                    type=DIVIDEND_MOVEMENTS[movement],
                    quantity=quantity,
                    gross_value_per_share=unit_price,
                    payment_date=parsed_date,
                )
            )
        elif movement in CORPORATE_ACTION_MOVEMENTS:
            if quantity is None or quantity <= 0:
                review("Corporate action row without a readable quantity.")
                continue
            result.corporate_actions.append(
                ParsedCorporateActionEvent(
                    row=row_number,
                    ticker=ticker,
                    name=product_name,
                    type=CORPORATE_ACTION_MOVEMENTS[movement],
                    date=parsed_date,
                    direction=direction,  # type: ignore[arg-type]
                    quantity=quantity,
                )
            )
        else:
            review(f'Unrecognized movement type: "{cell("movement")}".')

    return result
