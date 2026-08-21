import io
from pathlib import Path

import pytest

from backend.infrastructure.b3_import.statement_parser import (
    StatementFormatError,
    parse_statement,
)

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
SAMPLE_CSV = (FIXTURES / "b3_movimentacao_sample.csv").read_bytes()


def _sample_as_xlsx() -> bytes:
    """Same statement content, but as the portal's Excel export."""
    from openpyxl import Workbook

    text = SAMPLE_CSV.decode("utf-8")
    workbook = Workbook()
    sheet = workbook.active
    for line in text.strip().splitlines():
        sheet.append(line.split(";"))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parses_buys_and_sells_from_liquidation_rows():
    statement = parse_statement(SAMPLE_CSV, "extrato.csv")
    trades = {(t.ticker, t.type): t for t in statement.trades}
    assert len(statement.trades) == 3

    buy = trades[("ITSA4", "buy")]
    assert buy.quantity == 100
    assert buy.price_per_share == 9.10
    assert buy.date == "2026-01-10"

    sell = trades[("ITSA4", "sell")]
    assert sell.quantity == 40
    assert sell.price_per_share == 10.0

    assert trades[("MXRF11", "buy")].date == "2026-02-10"


def test_parses_dividends_with_type_classification():
    statement = parse_statement(SAMPLE_CSV, "extrato.csv")
    by_type = {(d.ticker, d.type): d for d in statement.dividends}
    assert len(statement.dividends) == 3

    assert by_type[("ITSA4", "dividend")].gross_value_per_share == 0.25
    assert by_type[("ITSA4", "jcp")].gross_value_per_share == 0.10
    assert by_type[("MXRF11", "reit_income")].payment_date == "2026-02-15"


def test_parses_split_bonus_and_reverse_split_events():
    statement = parse_statement(SAMPLE_CSV, "extrato.csv")
    by_type = {(a.ticker, a.type): a for a in statement.corporate_actions}
    assert len(statement.corporate_actions) == 3

    split = by_type[("ITSA4", "split")]
    assert split.quantity == 60
    assert split.direction == "credit"
    assert split.date == "2026-03-01"

    bonus = by_type[("ITSA4", "bonus_shares")]
    assert bonus.quantity == 12

    reverse = by_type[("MXRF11", "reverse_split")]
    assert reverse.quantity == 45
    assert reverse.direction == "debit"


def test_unknown_movements_go_to_review_and_irrelevant_ones_are_skipped():
    statement = parse_statement(SAMPLE_CSV, "extrato.csv")
    # "Atualização" is silently ignored; "Movimento Desconhecido" needs review.
    assert len(statement.review_rows) == 1
    assert "Movimento Desconhecido" in statement.review_rows[0].reason
    assert statement.review_rows[0].row == 12


def test_excel_file_parses_identically_to_csv():
    from_csv = parse_statement(SAMPLE_CSV, "extrato.csv")
    from_xlsx = parse_statement(_sample_as_xlsx(), "extrato.xlsx")
    assert [(t.ticker, t.type, t.quantity, t.price_per_share, t.date) for t in from_xlsx.trades] == [
        (t.ticker, t.type, t.quantity, t.price_per_share, t.date) for t in from_csv.trades
    ]
    assert len(from_xlsx.dividends) == len(from_csv.dividends)
    assert len(from_xlsx.corporate_actions) == len(from_csv.corporate_actions)


def test_brazilian_number_formats_are_parsed():
    csv_content = (
        "Entrada/Saída;Data;Movimentação;Produto;Instituição;Quantidade;Preço unitário;Valor da Operação\n"
        "Credito;10/01/2026;Transferência - Liquidação;VALE3 - VALE ON;X;1.000;R$ 1.234,56;R$ 1.234.560,00\n"
    ).encode("utf-8")
    statement = parse_statement(csv_content, "extrato.csv")
    assert statement.trades[0].quantity == 1000
    assert statement.trades[0].price_per_share == 1234.56


def test_a_file_that_is_not_a_statement_is_rejected():
    with pytest.raises(StatementFormatError):
        parse_statement(b"id,name\n1,foo\n", "random.csv")


def test_a_trade_row_without_a_price_goes_to_review():
    csv_content = (
        "Entrada/Saída;Data;Movimentação;Produto;Instituição;Quantidade;Preço unitário;Valor da Operação\n"
        "Credito;10/01/2026;Transferência - Liquidação;VALE3 - VALE ON;X;100;-;-\n"
    ).encode("utf-8")
    statement = parse_statement(csv_content, "extrato.csv")
    assert statement.trades == []
    assert len(statement.review_rows) == 1


def _sample_with(delimiter: str, preamble: list[str] | None = None) -> bytes:
    """Re-serializes the reference statement with another delimiter, optionally
    behind title/filter lines like some portal exports carry."""
    lines = SAMPLE_CSV.decode("utf-8").strip().splitlines()
    rebuilt = [delimiter.join(line.split(";")) for line in lines]
    return "\n".join((preamble or []) + rebuilt).encode("utf-8")


def test_tab_separated_export_parses_like_the_semicolon_one():
    reference = parse_statement(SAMPLE_CSV, "extrato.csv")
    tabbed = parse_statement(_sample_with("\t"), "extrato.csv")
    assert len(tabbed.trades) == len(reference.trades)
    assert len(tabbed.dividends) == len(reference.dividends)
    assert len(tabbed.corporate_actions) == len(reference.corporate_actions)
    assert tabbed.trades[0].price_per_share == reference.trades[0].price_per_share


def test_comma_separated_export_is_not_confused_by_decimal_commas():
    # Values keep the Brazilian decimal comma, so a naive frequency count would
    # pick "," even in a tab file — and split every price in half.
    tabbed = parse_statement(_sample_with("\t"), "extrato.csv")
    assert tabbed.trades[0].price_per_share == 9.10


def test_title_rows_above_the_header_are_skipped():
    statement = parse_statement(
        _sample_with(";", preamble=["Extrato de movimentação", "Período: 01/01/2026 a 31/12/2026", ""]),
        "extrato.csv",
    )
    assert len(statement.trades) == 3
    # Row numbers stay aligned with the actual file, for the review queue.
    assert statement.trades[0].row == 5


def test_the_error_lists_the_columns_it_did_find():
    content = b"Coluna A;Coluna B\nvalor;valor\n"
    with pytest.raises(StatementFormatError) as error:
        parse_statement(content, "extrato.csv")
    assert "Coluna A" in str(error.value)


def _workbook_bytes(sheets: dict[str, list[list]]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title=title)
        for row in rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


HEADER = [
    "Entrada/Saída", "Data", "Movimentação", "Produto",
    "Instituição", "Quantidade", "Preço unitário", "Valor da Operação",
]
TRADE_ROW = [
    "Credito", "10/01/2026", "Transferência - Liquidação",
    "ITSA4 - ITAUSA S.A.", "CORRETORA", 100, 9.10, 910.00,
]


def test_the_statement_is_found_even_when_it_is_not_the_first_sheet():
    content = _workbook_bytes(
        {
            "Capa": [["Extrato de movimentação"], ["Entrada/Saída"]],
            "Movimentação": [HEADER, TRADE_ROW],
        }
    )
    statement = parse_statement(content, "extrato.xlsx")
    assert len(statement.trades) == 1
    assert statement.trades[0].ticker == "ITSA4"
    assert statement.trades[0].price_per_share == 9.10


def test_a_cover_sheet_that_merely_mentions_a_column_name_does_not_win():
    # A single stray "Entrada/Saída" label must not beat the real header row —
    # this is exactly the shape that produced a confusing import failure.
    content = _workbook_bytes(
        {
            "Filtros": [["Entrada/Saída"], ["Período"]],
            "Dados": [HEADER, TRADE_ROW],
        }
    )
    statement = parse_statement(content, "extrato.xlsx")
    assert len(statement.trades) == 1


def test_plain_compra_and_venda_movements_are_recognized_as_trades():
    """Some brokers report the settlement as "Compra"/"Venda" rather than the
    portal's own "Transferência - Liquidação" — losing those loses real trades."""
    csv_content = (
        "Entrada/Saída;Data;Movimentação;Produto;Instituição;Quantidade;Preço unitário;Valor da Operação\n"
        "Credito;10/01/2026;Compra;POMO4 - MARCOPOLO PN;COR;100;R$ 7,00;R$ 700,00\n"
        "Debito;20/01/2026;Venda;POMO4 - MARCOPOLO PN;COR;40;R$ 8,00;R$ 320,00\n"
    ).encode("utf-8")
    statement = parse_statement(csv_content, "extrato.csv")
    assert [(t.type, t.quantity) for t in statement.trades] == [("buy", 100), ("sell", 40)]
    assert statement.review_rows == []


def test_subscription_lifecycle_rows_are_ignored_not_flagged():
    """Receiving, requesting or letting a subscription right lapse does not move
    the position — only an exercised right does (business-rules §5)."""
    header = (
        "Entrada/Saída;Data;Movimentação;Produto;Instituição;Quantidade;Preço unitário;Valor da Operação\n"
    )
    rows = [
        "Credito;10/01/2026;Direito de Subscrição;AFHI12 - AF INVEST;COR;5;-;-",
        "Credito;11/01/2026;Direito Sobras de Subscrição;AFHI12 - AF INVEST;COR;2;-;-",
        "Debito;12/01/2026;Direito Sobras de Subscrição - Não Exercido;AFHI12 - AF INVEST;COR;2;-;-",
        "Credito;13/01/2026;Recibo de Subscrição;AFHI13 - AF INVEST;COR;5;-;-",
        "Debito;14/01/2026;Solicitação de Subscrição;AFHI11 - AF INVEST;COR;5;-;-",
    ]
    statement = parse_statement((header + "\n".join(rows) + "\n").encode("utf-8"), "extrato.csv")
    assert statement.trades == []
    assert statement.review_rows == []


def test_an_exercised_subscription_right_is_still_a_buy():
    csv_content = (
        "Entrada/Saída;Data;Movimentação;Produto;Instituição;Quantidade;Preço unitário;Valor da Operação\n"
        "Credito;15/01/2026;Direitos de Subscrição - Exercido;AFHI11 - AF INVEST;COR;5;R$ 95,00;R$ 475,00\n"
    ).encode("utf-8")
    statement = parse_statement(csv_content, "extrato.csv")
    assert [(t.type, t.quantity, t.price_per_share) for t in statement.trades] == [("buy", 5, 95.0)]


def test_the_product_name_is_carried_for_asset_registration():
    statement = parse_statement(SAMPLE_CSV, "extrato.csv")
    assert statement.trades[0].name == "ITAUSA S.A."


def test_a_product_without_a_ticker_goes_to_review_instead_of_breaking_the_import():
    """A description with no ticker used to be taken whole as the ticker, which
    overflowed the 20-character column and aborted the whole import with a
    database error. It has to be a reviewable row, not a crash."""
    csv_content = (
        "Entrada/Saída;Data;Movimentação;Produto;Instituição;Quantidade;Preço unitário;Valor da Operação\n"
        "Credito;10/01/2026;Compra;FDO INV IMOB TRX REAL ESTATE LOG;COR;10;R$ 100,00;R$ 1.000,00\n"
    ).encode("utf-8")
    statement = parse_statement(csv_content, "extrato.csv")
    assert statement.trades == []
    assert len(statement.review_rows) == 1
    assert "Produto" in statement.review_rows[0].reason


def test_fixed_income_descriptions_are_not_mistaken_for_tickers():
    # Requiring the numeric suffix is what keeps "TESOURO SELIC 2029" out of the
    # catalog — fixed income is not supported by the import yet.
    from backend.infrastructure.b3_import.statement_parser import _extract_product

    assert _extract_product("TESOURO SELIC 2029")[0] is None
    assert _extract_product("CDB BANCO INTER")[0] is None


@pytest.mark.parametrize(
    "product,expected",
    [
        ("ITSA4 - ITAUSA S.A.", "ITSA4"),
        ("TAEE11 - TAESA", "TAEE11"),
        ("AFHI13 - AF INVEST", "AFHI13"),
        ("BOVA11 - ISHARES IBOVESPA", "BOVA11"),
        ("AAPL34 - APPLE BDR", "AAPL34"),
        ("ITSA4 ITAUSA S.A.", "ITSA4"),  # no " - " separator
        ("MXRF11", "MXRF11"),  # ticker only
    ],
)
def test_ticker_extraction_across_product_layouts(product, expected):
    from backend.infrastructure.b3_import.statement_parser import _extract_product

    assert _extract_product(product)[0] == expected


def test_no_extracted_ticker_can_overflow_the_database_column():
    """The assets table stores a 20-character ticker; the parser must never hand
    the import something longer."""
    from backend.infrastructure.b3_import.statement_parser import _extract_product

    long_products = [
        "FDO INV IMOB MUITO LONGO DEMAIS PARA CABER",
        "A" * 60,
        "TESOURO IPCA+ COM JUROS SEMESTRAIS 2045",
    ]
    for product in long_products:
        ticker, _ = _extract_product(product)
        assert ticker is None or len(ticker) <= 20
